#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Conciliador de Vendas -> Planilhas para o Sistema Financeiro
============================================================

Le os relatorios mensais de tres fontes (adquirente de cartoes, marketplace de
delivery e sistema de PDV/caixa) e gera as planilhas no layout de importacao do
sistema financeiro (ERP), uma por forma de pagamento, mais uma planilha de
divergencias.

O sistema de PDV e a fonte-mae (concentra todas as vendas). As outras fontes
servem de conferencia. Ver docs/ARQUITETURA.md e docs/MODELO_IMPORTACAO.md.

Uso:
    python gerar_planilhas.py <pasta_do_mes> <sufixo>

Exemplo:
    python gerar_planilhas.py ./dados/2026-05 mai2026

A <pasta_do_mes> deve conter os 3 relatorios (nomes reconhecidos por padrao):
    - Adquirente (cartoes):   Vendas_*cielo*detalhe*.xlsx
    - Marketplace (delivery): relatorios.ifood*.xlsx
    - PDV / caixa:            Relatorio_Geral_Vendas*.xlsx

Saida: subpasta SAIDA_Financeiro/ com os arquivos financeiro_<forma>_<sufixo>.xlsx
e DIVERGENCIAS_ifood_sem_nfce_<sufixo>.xlsx

Dependencias: openpyxl  (pip install openpyxl)
"""
import os
import sys
import glob
import unicodedata
import warnings
from datetime import datetime, timedelta
from collections import defaultdict, deque

import openpyxl
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# --------------------------------------------------------------------------- #
# Layout do modelo de importacao do sistema financeiro
# --------------------------------------------------------------------------- #
# Textos e cabecalhos EXATOS do modelo de importacao (com acentos) — o sistema
# financeiro casa as colunas pelo nome, entao precisam ser identicos.
ORIENTACOES = [
    'Orientações de preenchimento da planilha:',
    '* A data de pagamento precisa ser igual ou inferior a data de hoje, caso a '
    'mesma seja superior ao dia de hoje o lançamento será importado com o status: '
    '"Em Aberto".',
    '* Não utilizar caracteres especiais;',
    '* Cole as informações na planilha utilizando a função "Colar Especial > Colar '
    'Valores" para não perder a formatação padrão das células;',
    '* Verificar se não ficou espaços entre os dados informados, principalmente '
    'quando as informações são coladas;',
    '* As células não podem conter fórmulas;',
]
COLUNAS = [
    'Data de Competência', 'Data de Vencimento', 'Data de Pagamento', 'Valor',
    'Categoria', 'Descrição', 'Cliente/Fornecedor', 'CNPJ/CPF Cliente/Fornecedor',
    'Centro de Custo', 'Observações',
]


# --------------------------------------------------------------------------- #
# Helpers genericos
# --------------------------------------------------------------------------- #
def strip_accents(s):
    s = unicodedata.normalize('NFKD', str(s))
    return ''.join(c for c in s if not unicodedata.combining(c))


def norm(s):
    """Normaliza um rotulo de coluna: sem acento, minusculo, sem espacos extra."""
    return strip_accents(s).strip().lower()


def header_index(row):
    """Mapa {rotulo_normalizado: indice} para uma linha de cabecalho."""
    return {norm(c): i for i, c in enumerate(row) if c not in (None, '')}


def col(hmap, *names):
    """Indice da 1a coluna cujo rotulo casa (exato ou por prefixo, sem acento)."""
    for name in names:
        key = norm(name)
        if key in hmap:
            return hmap[key]
        for k, i in hmap.items():
            if k.startswith(key):
                return i
    raise KeyError(names)


def to_number(x):
    try:
        return round(float(str(x).replace(',', '.')), 2)
    except (TypeError, ValueError):
        return None


def parse_date(x):
    if x in (None, ''):
        return None
    s = str(x).split(' ')[0]
    for fmt in ('%d/%m/%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d'):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def fmt_date(d):
    return d.strftime('%d/%m/%Y') if d else ''


def find_file(folder, *patterns):
    for pattern in patterns:
        hits = glob.glob(os.path.join(folder, pattern))
        if hits:
            return hits[0]
    raise FileNotFoundError('Nenhum arquivo encontrado para: %s' % (patterns,))


def load_rows(path, must_have=None):
    """Le a planilha. Se `must_have` for dado, escolhe a aba cujo cabecalho
    contem esse rotulo (robusto a nome de aba com acento); senao usa a ativa."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    if must_have:
        alvo = norm(must_have)
        for cand in wb.worksheets:
            first = next(cand.iter_rows(values_only=True), ())
            if any(norm(c).startswith(alvo) for c in first if c not in (None, '')):
                ws = cand
                break
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def write_sheet(outdir, filename, rows):
    """Gera um arquivo no modelo de importacao (abas Orientacoes + Dados)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orientações'
    for line in ORIENTACOES:
        ws.append([line])
    dados = wb.create_sheet('Dados')
    dados.append(COLUNAS)
    for r in rows:
        dados.append(r + [''] * (len(COLUNAS) - len(r)))
    wb.save(os.path.join(outdir, filename))
    return len(rows)


# --------------------------------------------------------------------------- #
# Pipeline
# --------------------------------------------------------------------------- #
def gerar(folder, suf):
    outdir = os.path.join(folder, 'SAIDA_Financeiro')
    os.makedirs(outdir, exist_ok=True)

    # ===== Adquirente de cartoes -> Debito e Credito ======================== #
    # O export tem cabecalho institucional nas primeiras linhas: localizamos a
    # linha de header dinamicamente (nao fixamos indice).
    cartoes = load_rows(find_file(folder, 'Vendas_*cielo*detalhe*.xlsx', '*cielo*.xlsx'))
    hrow = next(i for i, r in enumerate(cartoes[:25])
                if any(norm(c) == 'data da venda' for c in r if c is not None))
    H = header_index(cartoes[hrow])

    debito, credito = [], []
    for r in cartoes[hrow + 1:]:
        if str(r[col(H, 'Status da venda')]) != 'Aprovada':
            continue
        forma = str(r[col(H, 'Forma de pagamento')])
        competencia = parse_date(r[col(H, 'Data da venda')])
        if competencia is None:
            continue
        vencimento = parse_date(r[col(H, 'Data prevista do pagamento')])
        bandeira = str(r[col(H, 'Bandeira')])
        valor = to_number(r[col(H, 'Valor bruto')])
        taxa = to_number(r[col(H, 'Taxa/tarifa')])
        nsu = str(r[col(H, 'NSU/DOC')])
        descricao = 'Taxa/Tarifa: R$ %s | NSU: %s' % (taxa, nsu)
        linha = [fmt_date(competencia), fmt_date(vencimento), '', valor,
                 bandeira, descricao, '', '', '', '']
        if 'bito' in forma:      # Debito
            debito.append(linha)
        elif 'dito' in forma:    # Credito
            credito.append(linha)

    # ===== PDV / caixa -> PIX, Dinheiro, Hanzo (+ pool para iFood) ========== #
    pdv = load_rows(find_file(folder, 'Relatorio_Geral_Vendas*.xlsx', 'Relatorio*Vendas*.xlsx'),
                    must_have='Data de Negocio')
    P = header_index(pdv[0])

    pix, dinheiro, hanzo = [], [], []
    pool_ifood = defaultdict(deque)  # (data, valor_bruto) -> [cupom fiscal, ...]

    for r in pdv[1:]:
        # Filtro global: apenas vendas efetivadas.
        if str(r[col(P, 'Status da Venda')]) != 'Pago':
            continue
        forma = str(r[col(P, 'Forma de pagamento')])
        competencia = parse_date(r[col(P, 'Data de Negocio')])
        if competencia is None:
            continue
        valor = to_number(r[col(P, 'Valor Pagamento')])
        if not valor or valor <= 0:          # ignora linhas zeradas/invalidas
            continue
        cupom = str(r[col(P, 'Cupom Fiscal')] or '').strip()   # numero da NFC-e
        pedido = str(r[col(P, 'Pedido')] or '').strip()
        venda_bruta = to_number(r[col(P, 'Venda Bruta')])

        if forma == 'PIX MANUAL':
            pix.append([fmt_date(competencia), fmt_date(competencia), '', valor,
                        'PIX - PIX', 'NFCe: %s | Pedido: %s' % (cupom, pedido),
                        '', '', '', ''])
        elif forma == 'Hanzo Prd':
            vencimento = competencia + timedelta(days=31)
            hanzo.append([fmt_date(competencia), fmt_date(vencimento), '', valor,
                          'Hanzo Prod', 'NFCe: %s | Pedido: %s' % (cupom, pedido),
                          '', '', '', ''])
        elif forma == 'Dinheiro':
            dinheiro.append([fmt_date(competencia), fmt_date(competencia), '', valor,
                             'Dinheiro', 'Pedido: %s | NFCe: %s' % (pedido, cupom),
                             '', '', '', ''])
        elif forma == 'iFood':
            # guarda a NFC-e do PDV indexada por (data, valor bruto) para casar
            # depois com o pedido do marketplace pelo valor.
            pool_ifood[(fmt_date(competencia), venda_bruta)].append(cupom)

    # ===== Marketplace de delivery -> iFood ================================ #
    # Casamento por VALOR (o numero do pedido do marketplace NAO e o numero do
    # PDV). A descricao une a NFC-e (PDV) com o ID do pedido (marketplace).
    mkt = load_rows(find_file(folder, 'relatorios.ifood*.xlsx', '*ifood*.xlsx'))
    M = header_index(mkt[0])
    m_id = col(M, 'ID CURTO')
    m_id_full = col(M, 'ID COMPLETO')
    m_datahora = col(M, 'DATA E HORA')
    m_status = col(M, 'STATUS FINAL')
    m_itens = col(M, 'VALOR DOS ITENS')
    m_total = col(M, 'TOTAL PAGO')
    m_liquido = col(M, 'VALOR LIQUIDO')
    m_forma = col(M, 'FORMA DE PAGAMENTO')

    ifood, divergencias = [], []
    for r in mkt[1:]:
        if str(r[m_status]) != 'CONCLUIDO':
            continue
        competencia = parse_date(r[m_datahora])
        if competencia is None:
            continue
        valor = to_number(r[m_itens])
        id_curto = str(r[m_id])
        chave = (fmt_date(competencia), valor)
        cupom = pool_ifood[chave].popleft() if pool_ifood.get(chave) else ''
        if cupom:
            descricao = 'NFCe: %s | ID iFood: %s' % (cupom, id_curto)
        else:
            descricao = 'ID iFood: %s' % id_curto
        ifood.append([fmt_date(competencia), fmt_date(competencia), '', valor,
                      'Ifood', descricao, '', '', '', ''])
        if not cupom:  # pedido do marketplace sem NFC-e correspondente no PDV
            datahora = str(r[m_datahora])
            hora = datahora.split(' ')[1] if ' ' in datahora else ''
            divergencias.append([fmt_date(competencia), hora, id_curto,
                                  str(r[m_id_full]), valor, to_number(r[m_total]),
                                  to_number(r[m_liquido]), str(r[m_forma])])

    # ===== Escreve as saidas =============================================== #
    formas = [
        ('Debito',   'financeiro_debito_%s.xlsx' % suf,   debito),
        ('Credito',  'financeiro_credito_%s.xlsx' % suf,  credito),
        ('PIX',      'financeiro_pix_%s.xlsx' % suf,      pix),
        ('Dinheiro', 'financeiro_dinheiro_%s.xlsx' % suf, dinheiro),
        ('Hanzo',    'financeiro_hanzo_%s.xlsx' % suf,    hanzo),
        ('iFood',    'financeiro_ifood_%s.xlsx' % suf,    ifood),
    ]
    resumo_formas = []
    for nome, arquivo, linhas in formas:
        write_sheet(outdir, arquivo, linhas)
        total = round(sum(x[3] for x in linhas if x[3]), 2)
        resumo_formas.append({'forma': nome, 'arquivo': arquivo,
                              'linhas': len(linhas), 'total': total})

    # Planilha de divergencias (pedidos do marketplace sem NFC-e no PDV)
    div_arquivo = 'DIVERGENCIAS_ifood_sem_nfce_%s.xlsx' % suf
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'iFood sem NFCe no PDV'
    cols = ['Data', 'Hora', 'ID Curto iFood', 'ID Completo iFood', 'Valor dos Itens',
            'Total Pago Cliente', 'Valor Liquido', 'Forma de Pagamento']
    ws.append(cols)
    for d in sorted(divergencias, key=lambda x: (x[0], x[1])):
        ws.append(d)
    for i, name in enumerate(cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(name) + 2)
    wb.save(os.path.join(outdir, div_arquivo))

    # Resumo da execucao (usado pela interface web e pelo CLI).
    return {
        'saida': outdir,
        'formas': resumo_formas,
        'divergencias': {'arquivo': div_arquivo, 'linhas': len(divergencias)},
        'linhas_total': sum(f['linhas'] for f in resumo_formas),
        'valor_total': round(sum(f['total'] for f in resumo_formas), 2),
    }


def main():
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    resultado = gerar(sys.argv[1], sys.argv[2])
    print('Planilhas geradas em:', resultado['saida'])


if __name__ == '__main__':
    main()
